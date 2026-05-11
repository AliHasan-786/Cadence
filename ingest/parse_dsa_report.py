"""Parse Spotify DSA XLSX annexes into validated Parquet.

For each of the 4 product lines × 9 sheets, this:
    1. Reads the XLSX from precache/dsa_reports/raw/{product}_{period}.xlsx
    2. Verifies the column header matches the harmonised template exactly
    3. Validates every non-empty row through a pydantic schema
    4. Writes one Parquet file per (product × sheet) to
       precache/dsa_reports/parquet/{product}_{period}/{sheet_slug}.parquet

Schema-drift safety: any header diff vs. the expected harmonised template
aborts with a precise diff report. Any row that fails pydantic validation
aborts with the row index and the validation error.

Usage:
    uv run python -m ingest.parse_dsa_report --period 2025H2
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from pydantic import ValidationError

from ingest.schemas.dsa_harmonised_v2 import SHEET_REGISTRY, _norm

# Field-name patterns identifying numeric columns. Drives explicit pyarrow types
# at Parquet write time so all 4 products produce identical schemas — otherwise
# pandas infers `null` dtype for products where the column is all-NULL and
# `double` for products where Spotify happens to write a sentinel 0.
_NUMERIC_PREFIXES = (
    "n_",
    "median_",
    "vis_restriction_",
    "mon_restriction_",
    "provision_",
    "account_",
)
# In sheets 7 (appeals) and 8 (automated_means), the catch-all `value` column
# holds numeric counts/rates. In sheets 1 (report_identification) and 9
# (qualitative) the same field name holds text. Class-level override.
_NUMERIC_VALUE_CLASSES = {"AppealsAndRecidivismRow", "AutomatedMeansRow"}


def _pa_type_for_field(class_name: str, field_name: str) -> pa.DataType:
    if field_name == "source_row_index":
        return pa.int64()
    if field_name.startswith("source_") or field_name.startswith("ctx_"):
        return pa.string()
    if any(field_name.startswith(p) for p in _NUMERIC_PREFIXES):
        return pa.float64()
    if field_name == "value" and class_name in _NUMERIC_VALUE_CLASSES:
        return pa.float64()
    return pa.string()


def _build_pa_schema(class_name: str, field_names: list[str]) -> pa.Schema:
    return pa.schema([pa.field(name, _pa_type_for_field(class_name, name)) for name in field_names])


def _coerce_dataframe(df: pd.DataFrame, schema: pa.Schema) -> pd.DataFrame:
    """Coerce DataFrame columns to match the pyarrow schema's types."""
    out = pd.DataFrame()
    for field in schema:
        col = df[field.name] if field.name in df.columns else pd.Series([None] * len(df))
        if pa.types.is_integer(field.type):
            out[field.name] = pd.to_numeric(col, errors="coerce").astype("Int64")
        elif pa.types.is_floating(field.type):
            out[field.name] = pd.to_numeric(col, errors="coerce")
        else:
            out[field.name] = col.apply(
                lambda v: None if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)
            )
    return out


REPO_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = REPO_ROOT / "precache" / "dsa_reports" / "raw"
PARQUET_DIR = REPO_ROOT / "precache" / "dsa_reports" / "parquet"

PRODUCTS = ("main", "artists", "authors", "creators")


@dataclass
class SheetResult:
    product: str
    period: str
    sheet_name: str
    sheet_slug: str
    n_rows_in_sheet: int
    n_rows_written: int
    parquet_path: Path


def _row_has_data(row: tuple[Any, ...]) -> bool:
    """A row 'has data' iff at least one cell is non-empty after normalisation."""
    for v in row:
        if _norm(v) is not None:
            return True
    return False


def _parse_sheet(
    *,
    workbook: openpyxl.Workbook,
    sheet_name: str,
    product: str,
    period: str,
    source_sha256: str,
) -> SheetResult:
    slug, RowSchema, expected_headers = SHEET_REGISTRY[sheet_name]  # noqa: N806 RowSchema is a class reference
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError(f"{product}/{period}/{sheet_name}: empty sheet")

    # Trim raw header to the canonical width
    actual_header = tuple(h for h in rows[0][: len(expected_headers)])
    if actual_header != expected_headers:
        diff_lines = []
        for i, (got, want) in enumerate(zip(actual_header, expected_headers, strict=False)):
            if got != want:
                diff_lines.append(f"  col {i} ({chr(65 + i)}): got {got!r}  want {want!r}")
        raise RuntimeError(
            f"{product}/{period}/{sheet_name}: header drift detected\n"
            + ("\n".join(diff_lines) if diff_lines else "(width mismatch)")
            + f"\n  actual width: {len(actual_header)}; expected: {len(expected_headers)}"
        )

    validated_rows: list[dict[str, Any]] = []
    width = len(expected_headers)

    for excel_row_idx, raw_row in enumerate(rows[1:], start=2):  # row 1 is header
        # Truncate to canonical width, normalise None/empty/n/a, skip if entirely empty
        cells = tuple(_norm(raw_row[i]) if i < len(raw_row) else None for i in range(width))
        if not _row_has_data(cells):
            continue
        kwargs: dict[str, Any] = dict(zip(expected_headers, cells, strict=True))
        kwargs.update(
            source_product=product,
            source_period=period,
            source_sheet=sheet_name,
            source_row_index=excel_row_idx,
            source_sha256=source_sha256,
        )
        try:
            obj = RowSchema.model_validate(kwargs)
        except ValidationError as e:
            raise RuntimeError(
                f"{product}/{period}/{sheet_name} row {excel_row_idx}: "
                f"validation failed\n{e}\n  raw row: {raw_row}"
            ) from e
        validated_rows.append(obj.model_dump(by_alias=False))

    # Write Parquet
    out_dir = PARQUET_DIR / f"{product}_{period}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{slug}.parquet"

    # Build the explicit pyarrow schema from the pydantic model field order
    field_names = list(RowSchema.model_fields.keys())
    pa_schema = _build_pa_schema(RowSchema.__name__, field_names)

    if validated_rows:
        df = pd.DataFrame(validated_rows)
        df = _coerce_dataframe(df, pa_schema)
        table = pa.Table.from_pandas(df, schema=pa_schema, preserve_index=False)
    else:
        # Schema-only Parquet so downstream loads still find a file
        table = pa.Table.from_arrays(
            [pa.array([], type=f.type) for f in pa_schema],
            schema=pa_schema,
        )

    pq.write_table(table, out_path, compression="snappy")

    return SheetResult(
        product=product,
        period=period,
        sheet_name=sheet_name,
        sheet_slug=slug,
        n_rows_in_sheet=len(rows) - 1,
        n_rows_written=len(validated_rows),
        parquet_path=out_path,
    )


def _load_sha256(product: str, period: str) -> str:
    meta = RAW_DIR / f"{product}_{period}.meta.json"
    if not meta.exists():
        raise FileNotFoundError(
            f"{meta} missing — run `ingest.fetch_spotify_dsa --period {period}` first"
        )
    return json.loads(meta.read_text())["sha256"]


def parse_product(product: str, period: str) -> list[SheetResult]:
    xlsx_path = RAW_DIR / f"{product}_{period}.xlsx"
    if not xlsx_path.exists():
        raise FileNotFoundError(f"{xlsx_path} missing — run fetch_spotify_dsa first")
    sha = _load_sha256(product, period)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    results: list[SheetResult] = []
    for sheet_name in SHEET_REGISTRY:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(
                f"{product}/{period}: expected sheet {sheet_name!r} missing. "
                f"Available: {wb.sheetnames}"
            )
        result = _parse_sheet(
            workbook=wb,
            sheet_name=sheet_name,
            product=product,
            period=period,
            source_sha256=sha,
        )
        results.append(result)
    wb.close()
    return results


def parse_period(period: str) -> dict[str, list[SheetResult]]:
    all_results: dict[str, list[SheetResult]] = {}
    for product in PRODUCTS:
        print(f"\n=== {product}/{period} ===")
        results = parse_product(product, period)
        for r in results:
            print(
                f"  {r.sheet_slug:30s} rows_in_sheet={r.n_rows_in_sheet:>5}  "
                f"rows_written={r.n_rows_written:>5}  → {r.parquet_path.relative_to(REPO_ROOT)}"
            )
        all_results[product] = results
    return all_results


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--period", default="2025H2")
    args = parser.parse_args(argv)
    all_results = parse_period(args.period)

    # Summary
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    total_rows = 0
    total_files = 0
    for product, results in all_results.items():
        product_rows = sum(r.n_rows_written for r in results)
        print(f"  {product:10s}: {len(results)} sheets, {product_rows} validated rows")
        total_rows += product_rows
        total_files += len(results)
    print(f"\nTotal: {total_files} Parquet files, {total_rows} validated rows")
    print(f"Output: {PARQUET_DIR}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
